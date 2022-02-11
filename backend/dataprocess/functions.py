import datetime
from gc import collect
from dateutil.parser import parse
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, fonts
from openpyxl.styles.alignment import Alignment
from config.models import PlatformTargetItem, CollectTargetItem, Schedule, AuthInfo
from dataprocess.models import Platform, Artist, CollectTarget, CollectData

from config.serializers import PlatformTargetItemSerializer, CollectTargetItemSerializer, ScheduleSerializer, AuthInfoSerializer
from dataprocess.serializers import ArtistSerializer, PlatformSerializer, CollectTargetSerializer

def get_platform_data(artist, platform, type, start_date, end_date, collect_item_list):
    filter_datas = []
    # 길이 초기화
    for i in collect_item_list:
        filter_datas.append("NULL")

    # 오늘 날짜 기준으로 가져오기
    if type == "누적":
        # today_date = datetime.datetime.today()
        start_date_dateobject = datetime.datetime.strptime(start_date, "%Y-%m-%d")
        start_date_string = start_date_dateobject.strftime("%Y-%m-%d")
        filter_objects = CollectData.objects.filter(collect_items__artist=artist, collect_items__platform=platform,
            collect_items__reserved_date = start_date_string)
        if filter_objects.exists():
            filter_value = filter_objects.values().first()
            filter_value = filter_value["collect_items"]
            # 숫자필드값+user_created만 보내주기
            for field_name in filter_value.keys():
                if field_name != "id" and field_name != "artist" and field_name != "recorded_date" and field_name != "updated_at" and field_name != "reserved_date" and field_name != "platform" and field_name != "url" and field_name != "url1" and field_name != "url2" and field_name != "fans":
                    # 싱크 맞춰서 넣기
                    filter_datas[collect_item_list.index(field_name)] = filter_value[field_name]
            return filter_datas
        else:
            return filter_datas
    elif type == "기간별":
        start_date_dateobject = datetime.datetime.strptime(start_date, "%Y-%m-%d").date() - datetime.timedelta(1)
        end_date_dateobject = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
        start_date_string = start_date_dateobject.strftime("%Y-%m-%d")
        end_date_string = end_date_dateobject.strftime("%Y-%m-%d")
        filter_start_objects = CollectData.objects.filter(collect_items__artist=artist, collect_items__platform=platform,
            collect_items__reserved_date = start_date_string)
        filter_end_objects = CollectData.objects.filter(collect_items__artist=artist, collect_items__platform=platform,
            collect_items__reserved_date = end_date_string)
        if filter_start_objects.exists() and filter_end_objects.exists():
            filter_start_value = filter_start_objects.values().first()
            filter_end_value = filter_end_objects.values().first()
            filter_start_value = filter_start_value["collect_items"]
            filter_end_value = filter_end_value["collect_items"]
            # 숫자필드값+user_created만 보내주기
            for field_name in filter_start_value.keys():
                if field_name == "user_created":
                    filter_datas[collect_item_list.index(field_name)] = filter_start_value[field_name]
                elif field_name != "id" and field_name != "artist" and field_name != "recorded_date" and field_name != "updated_at" and field_name != "reserved_date" and field_name != "platform" and field_name != "url" and field_name != "url1" and field_name != "url2" and field_name != "fans":
                    if filter_end_value[field_name] is not None and filter_start_value[field_name] is not None:
                        filter_datas[collect_item_list.index(field_name)] = filter_end_value[field_name]-filter_start_value[field_name]
                        # 둘 중 하나라도 field값이 없으면 NULL로 들어감
            return filter_datas
        else:
            return filter_datas
    else:
        return filter_datas


def export_datareport(excel_export_type, excel_export_start_date, excel_export_end_date):
    # DB에서 platform과 platform_target_item 가져오기
    db_platform_datas = []
    platforms = Platform.objects.all()
    if platforms.exists():
        platform_objects_values = platforms.values()
        for platform_value in platform_objects_values:
            collecttargets = CollectTarget.objects.filter(platform=platform_value["id"])
            collecttargets = collecttargets.values()
            collect_item = set()
            for collecttarget in collecttargets:
                platform_objects = CollectTargetItem.objects.filter(collect_target_id=collecttarget["id"])
                platform_objects_values = platform_objects.values()
                for p in platform_objects_values:
                    if p["target_name"] in collect_item:
                        continue
                    collect_item.add(p["target_name"])
            collect_item = list(collect_item)

            platform_header = []
            objects = CollectData.objects.filter(collect_items__platform=platform_value["name"])
            objects_values = objects.values()
            if len(objects_values) > 0:
                key_list = list(objects_values[0]["collect_items"].keys())
                for key in key_list:
                    if key in collect_item:
                        platform_header.append(key)
                    else:
                        continue
            else:
                platform_header = collect_item
            db_platform_datas.append({
                "platform": platform_value["name"],
                "collect_item": platform_header
            })

    # DB에서 artist 가져오기
    db_artists = []
    artists = Artist.objects.all()
    if artists.exists():
        artist_objects_values = artists.values()
        for artist_value in artist_objects_values:
            db_artists.append(artist_value["name"])

    # export excel
    book = openpyxl.Workbook()
    sheet = book.active

    row = 1
    col = 1

    thick_border = Border(left=Side(style="medium"),
                          right=Side(style="medium"),
                          top=Side(style="medium"),
                          bottom=Side(style="medium"))

    sheet.cell(row=1, column=1).value = "플랫폼"
    sheet.cell(row=1, column=1).border = thick_border
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row=1, column=1).font = fonts.Font(bold=True)

    sheet.cell(row=2, column=1).value = "아티스트"
    sheet.cell(row=2, column=1).border = thick_border
    sheet.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    col += 1
    sheet.row_dimensions[2].height = 35
    sheet.column_dimensions["A"].width = 40

    # 플랫폼 이름과 수집항목 띄우기
    for data in db_platform_datas:
        if len(data["collect_item"]) == 0:
            continue
        sheet.cell(row=row, column=col).value = data["platform"]
        sheet.merge_cells(start_row=row, start_column=col,
                          end_row=row, end_column=col+len(data["collect_item"])-1)
        sheet.cell(row=row, column=col).border = thick_border
        sheet.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row, column=col).font = fonts.Font(bold=True)

        for i, collect_data in enumerate(data["collect_item"]):
            sheet.cell(row=row+1, column=col+i).value = collect_data
            sheet.cell(row=row+1, column=col+i).border = thick_border
            sheet.cell(row=row+1, column=col+i).alignment = Alignment(horizontal="center", vertical="center")
        col += len(data["collect_item"])

    # 아티스트별로 플랫폼순서대로 가져와서 띄우기
    row = 3
    col = 1
    for artist in db_artists:
        artist_name = artist
        sheet.cell(row=row, column=col).value = artist_name
        sheet.cell(row=row, column=col).border = Border(right=Side(style="medium"))
        col += 1
        for platform in db_platform_datas:
            # 아티스트의 플랫폼마다의 정보 가져오기
            platform_name = platform["platform"]
            platform_data_list = get_platform_data(artist=artist_name, platform=platform_name,
                                                   type=excel_export_type, start_date=excel_export_start_date, end_date=excel_export_end_date,
                                                   collect_item_list=platform["collect_item"])
            for i, platform_data in enumerate(platform_data_list):
                if platform_data is None or platform_data == "NULL":
                    # null이면 shade 처리
                    sheet.cell(row=row, column=col+i).fill = PatternFill(start_color="C4C4C4", end_color="C4C4C4", fill_type="solid")
                else:
                    sheet.cell(row=row, column=col+i).value = platform_data
            sheet.cell(row=row, column=col+len(platform["collect_item"])-1).border = Border(right=Side(style="medium"))
            col += len(platform["collect_item"])
        row += 1
        col = 1
    return book


def import_datareport(worksheet, excel_import_date):
    platform_data_list = []
    row_num = 0

    if excel_import_date is None:
        target_date = datetime.datetime.today()
    else:
        target_date = datetime.datetime.strptime(excel_import_date, "%Y-%m-%d")
    target_date = datetime.date(target_date.year, target_date.month, target_date.day)
    target_date_string = target_date.strftime("%Y-%m-%d")
    for row in worksheet.iter_rows():
        # 플랫폼 정보 나열
        if row_num == 0:
            item_num = 0
            platform_name = "platform"
            data_json = {}
            for cell in row:
                platform_value = str(cell.value)
                if platform_value == "플랫폼":
                    platform_name = "플랫폼"
                elif platform_value != "None":
                    if platform_name != "플랫폼":
                        # 새로운 게 등장한 거므로 저장
                        data_json["platform"] = platform_name
                        data_json["item_num"] = item_num
                        data_json["item_list"] = []
                        platform_data_list.append(data_json)
                        data_json = {}
                    item_num = 1
                    platform_name = platform_value
                else:  # None일때
                    item_num += 1
            # 마지막 저장
            data_json["platform"] = platform_name
            data_json["item_num"] = item_num
            data_json["item_list"] = []
            platform_data_list.append(data_json)
            row_num += 1
        # 아티스트 정보 나열
        elif row_num == 1:
            platform_index = 0
            current_index = 0
            for cell in row:
                collect_value = str(cell.value)
                if collect_value != "아티스트":
                    platform_data_list[platform_index]["item_list"].append(collect_value)
                    current_index += 1
                    if current_index >= platform_data_list[platform_index]["item_num"]:
                        platform_index += 1
                        current_index = 0
            row_num += 1
        # 데이터 정보 나열
        else:
            platform_index = 0
            current_index = 0
            artist_name = "artist"
            data_json = {}
            for i, cell in enumerate(row):
                # 아티스트 이름 나열
                if i == 0:
                    artist_name = str(cell.value)
                else:
                    # 아티스트와 플랫폼 이름에 대해 업데이트
                    platform_name = platform_data_list[platform_index]["platform"]
                    collect_value = platform_data_list[platform_index]["item_list"][current_index]
                    value = str(cell.value)
                    if value != "None":
                        # -가 있으면 날짜
                        if "-" in value:
                            value = str(cell.value)
                        # ,가 있으면 날짜
                        elif "," in value:
                            dateobject = parse(str(cell.value))
                            value = "%s-%s-%s"%(dateobject.year, dateobject.month, dateobject.day)
                        else:
                            value = int(cell.value)
                        data_json[collect_value] = value
                    current_index += 1
                    if current_index >= platform_data_list[platform_index]["item_num"]:
                        # 데이터 저장
                        data_json["artist"] = artist_name
                        save_collect_data_target(data_json, platform_name, target_date_string)
                        platform_index += 1
                        current_index = 0
                        data_json = {}


def import_collects(worksheet):
    platform_data_list = []
    artist_data_list = []
    collect_target_data_list = []
    collect_target_item_data_list = []
    platform_start_index = 7
    row_num = 0
    for row in worksheet.iter_rows():
        if row_num == 0 or row_num == 3 or row_num == 4:
            row_num += 1
            continue
        # 플랫폼 정보 나열
        if row_num == 1:
            item_num = 0
            platform_name = "플랫폼"
            data_json = {}
            for i, cell in enumerate(row):
                if i < platform_start_index:
                    continue
                platform_value = str(cell.value)
                if platform_value != "None":
                    if platform_name != "플랫폼":
                        # 새로운 게 등장한 거므로 저장
                        data_json["platform"] = platform_name
                        data_json["item_num"] = item_num
                        data_json["item_list"] = []
                        data_json["item_xpath_list"] = []
                        platform_data_list.append(data_json)
                        data_json = {}
                    item_num = 1
                    platform_name = platform_value
                else:  # None일때
                    item_num += 1
            # 마지막 저장
            data_json["platform"] = platform_name
            data_json["item_num"] = item_num
            data_json["item_list"] = []
            data_json["item_xpath_list"] = []
            platform_data_list.append(data_json)
            row_num += 1
        # 플랫폼 주소 나열
        elif row_num == 2:
            platform_index = 0
            current_index = 0
            for i, cell in enumerate(row):
                if i < platform_start_index:
                    continue
                collect_value = str(cell.value)
                current_index += 1
                if collect_value != "None":
                    platform_data_list[platform_index]["url"] = collect_value
                if current_index >= platform_data_list[platform_index]["item_num"]:
                    platform_index += 1
                    current_index = 0
            row_num += 1
        # 지표 정보(영어) 나열
        elif row_num == 5:
            platform_index = 0
            current_index = 0
            for i, cell in enumerate(row):
                if i < platform_start_index:
                    continue
                collect_value = str(cell.value)
                if collect_value != "지표 이름":
                    platform_data_list[platform_index]["item_list"].append(collect_value)
                    current_index += 1
                    if current_index >= platform_data_list[platform_index]["item_num"]:
                        platform_index += 1
                        current_index = 0
            row_num += 1
        # # 지표 xpath 정보 나열
        # elif row_num == 6:
        #     platform_index = 0
        #     current_index = 0
        #     for i, cell in enumerate(row):
        #         if i < platform_start_index:
        #             continue
        #         collect_value = str(cell.value)
        #         if collect_value != "xpath":
        #             platform_data_list[platform_index]["item_xpath_list"].append(collect_value)
        #             current_index += 1
        #             if current_index >= platform_data_list[platform_index]["item_num"]:
        #                 platform_index += 1
        #                 current_index = 0
        #     row_num += 1
        # 아티스트 & target_url 정보 나열
        else:
            platform_index = 0
            current_index = 0
            data_json = {}
            data_json2 = {}
            data_json3 = {}
            artist_name = ""
            for i, cell in enumerate(row):
                # 아티스트 이름 나열
                if i == 0:  # 이름
                    if str(cell.value) == "None":
                        break
                    data_json["name"] = str(cell.value)
                    artist_name = str(cell.value)
                elif i == 1:  # 구분
                    if str(cell.value) != "None":
                        data_json["level"] = str(cell.value)
                elif i == 2:  # 성별
                    if str(cell.value) != "None":
                        data_json["level"] = str(cell.value)
                elif i == 3:  # 멤버수
                    if str(cell.value) != "None":
                        data_json["member_num"] = int(cell.value)
                elif i == 4:  # 국적
                    if str(cell.value) != "None":
                        data_json["member_nationality"] = str(cell.value)
                elif i == 5:  # 기획사
                    if str(cell.value) != "None":
                        data_json["agency"] = str(cell.value)
                elif i == 6:  # 데뷔일
                    if str(cell.value) != "None":
                        strings = str(cell.value).split(" ")
                        data_json["debut_date"] = strings[0]
                    artist_data_list.append(data_json)
                    data_json = {}
                # target_url나올 때
                else:
                    collect_value = str(cell.value)
                    current_index += 1
                    if "http" in collect_value:
                        if "target_url" in data_json2:
                            data_json2["target_url_2"] = collect_value
                        else:
                            data_json2["target_url"] = collect_value
                    elif collect_value != " " and collect_value != "None" and platform_data_list[platform_index]["item_list"][current_index-1] != "None":
                        data_json3 = {
                            "platform": platform_data_list[platform_index]["platform"],
                            "artist": artist_name,
                            "target_name": platform_data_list[platform_index]["item_list"][current_index-1],
                            "xpath": collect_value
                        }
                        collect_target_item_data_list.append(data_json3)
                    if current_index >= platform_data_list[platform_index]["item_num"]:
                        platform_index += 1
                        current_index = 0
                        collect_target_data_list.append(data_json2)
                        data_json2 = {}

    # platform 저장
    for platform_data in platform_data_list:
        save_platform({
            "name": platform_data["platform"],
            "url": platform_data["url"]
        })
        # #platform_target_item 저장
        # for j in range(platform_data["item_num"]):
        #     if platform_data["item_list"][j] != "None" and platform_data["item_list"][j] != "url" and platform_data["item_list"][j] != "url1"and platform_data["item_list"][j] != "url2"and platform_data["item_xpath_list"][j] != "수집불가":
        #         platform_filter_object = Platform.objects.filter(name = platform_data["platform"])
        #         if platform_filter_object.exists():
        #             platform_filter_object = platform_filter_object.values().first()
        #             save_platform_target_item({
        #                 "platform": platform_filter_object["id"],
        #                 "target_name": platform_data["item_list"][j],
        #                 "xpath": platform_data["item_xpath_list"][j],
        #             })
    # artist 저장
    collect_target_index = 0
    for artist_data in artist_data_list:
        save_artist(artist_data)
        # collect_target 저장
        for platform_data in platform_data_list:
            if "target_url" in collect_target_data_list[collect_target_index]:
                platform_filter_object = Platform.objects.filter(name=platform_data["platform"])
                platform_filter_object = platform_filter_object.values().first()
                artist_filter_object = Artist.objects.filter(name=artist_data["name"])
                artist_filter_object = artist_filter_object.values().first()
                if 'target_url_2' in collect_target_data_list[collect_target_index]:
                    save_collect_target({
                    "artist": artist_filter_object['id'],
                    "platform": platform_filter_object['id'],
                    "target_url": collect_target_data_list[collect_target_index]['target_url'],
                    "target_url_2": collect_target_data_list[collect_target_index]['target_url_2'],
                    })
                else:
                    save_collect_target({"artist": artist_filter_object["id"],
                                         "platform": platform_filter_object["id"],
                                         "target_url": collect_target_data_list[collect_target_index]["target_url"]
                    })
                collecttarget_object = CollectTarget.objects.filter(artist_id = artist_filter_object['id'],
                    platform=platform_filter_object['id'])
                collecttarget_object = collecttarget_object.values().first()
                collecttargetid = collecttarget_object["id"]
                save_schedule(collecttargetid)
            collect_target_index += 1
    # collecttargetitem 저장
    for collect_target_item_data in collect_target_item_data_list:
        save_collect_target_item(collect_target_item_data)

def import_authinfo(worksheet):
    platform_data_list = []
    artist_data_list = []
    auth_info_list = []
    platform_start_index = 1
    row_num = 0
    for row in worksheet.iter_rows():
        if row_num == 0 or row_num == 3 or row_num == 4:
            row_num += 1
            continue
        # 플랫폼 정보 나열
        if row_num == 1:
            item_num = 0
            platform_name = "플랫폼"
            data_json = {}
            for i, cell in enumerate(row):
                if i < platform_start_index:
                    continue
                platform_value = str(cell.value)
                if platform_value != "None":
                    if platform_name != "플랫폼":
                        # 새로운 게 등장한 거므로 저장
                        data_json["platform"] = platform_name
                        data_json["item_num"] = item_num
                        data_json["item_list"] = []
                        platform_data_list.append(data_json)
                        data_json = {}
                    item_num = 1
                    platform_name = platform_value
                else:  # None일때
                    item_num += 1
            # 마지막 저장
            data_json["platform"] = platform_name
            data_json["item_num"] = item_num
            data_json["item_list"] = []
            platform_data_list.append(data_json)
            row_num += 1
        # 플랫폼 주소 나열
        elif row_num == 2:
            platform_index = 0
            current_index = 0
            for i, cell in enumerate(row):
                if i < platform_start_index:
                    continue
                collect_value = str(cell.value)
                current_index += 1
                if collect_value != "None":
                    platform_data_list[platform_index]["url"] = collect_value
                if current_index >= platform_data_list[platform_index]["item_num"]:
                    platform_index += 1
                    current_index = 0
            row_num += 1
        # 지표 정보(영어) 나열
        elif row_num == 5:
            platform_index = 0
            current_index = 0
            for i, cell in enumerate(row):
                if i < platform_start_index:
                    continue
                collect_value = str(cell.value)
                if collect_value != "지표 이름":
                    platform_data_list[platform_index]["item_list"].append(collect_value)
                    current_index += 1
                    if current_index >= platform_data_list[platform_index]["item_num"]:
                        platform_index += 1
                        current_index = 0
            row_num += 1
        # 아티스트 & auth_info 정보 나열
        else:
            platform_index = 0
            current_index = 0
            data_json = {}
            data_json2 = {}
            artist_name = ""
            for i, cell in enumerate(row):
                # 아티스트 이름 나열
                if i == 0:  # 이름
                    if str(cell.value) == "None":
                        break
                    data_json["name"] = str(cell.value)
                    artist_name = str(cell.value)
                    artist_data_list.append(artist_name)
                else: #auth_info
                    collect_value = str(cell.value)
                    if collect_value != 'null' and collect_value != 'None' and collect_value != ' ':
                        data_json2[platform_data_list[platform_index]["item_list"][current_index]] = collect_value
                    current_index += 1
                    if current_index >= platform_data_list[platform_index]["item_num"]:
                        if data_json2 != {}:
                            data_json2["artist"] = artist_name
                            data_json2["platform"] = platform_data_list[platform_index]["platform"]
                            auth_info_list.append(data_json2)
                            data_json2 = {}
                        platform_index += 1
                        current_index = 0
    # collecttargetitem 저장
    for auth_info_item in auth_info_list:
        save_auth_info(auth_info_item)

def save_auth_info(data_json):
    """
    수집(크롤링) 데이터 저장
    """
    platform = data_json["platform"]
    artist = data_json["artist"]
    artist_object = Artist.objects.filter(name = artist).first()
    platform_object = Platform.objects.filter(name = platform).first()
    collect_target_object = CollectTarget.objects.filter(platform_id = platform_object.id, artist_id = artist_object.id).first()
    new_data = {'collect_target': collect_target_object.id}
    for key in data_json.keys():
        if key == 'platform' or key == 'artist':
            continue
        new_data[key] = data_json[key]
    obj = AuthInfo.objects.filter(collect_target=collect_target_object.id).first()
    authinfo_serializer = AuthInfoSerializer(obj, data=new_data)
    if authinfo_serializer.is_valid():
        authinfo_serializer.save()
    else:
        print("===invalid===")


def save_collect_data_target(data_json, platform, target_date_string):
    """
    수집(크롤링) 데이터 저장
    """
    if len(data_json.keys())==1:
        return
    # collectdata 저장
    data_json["reserved_date"] = target_date_string
    data_json["platform"] = platform
    artist_object = Artist.objects.filter(name=data_json["artist"])
    artist_object = artist_object.values()[0]
    platform_object = Platform.objects.filter(name=platform)
    platform_object = platform_object.values()[0]
    collecttarget_object = CollectTarget.objects.filter(platform_id = platform_object["id"],
                            artist_id = artist_object["id"])
    collecttarget_object = collecttarget_object.values()[0]
    CollectData.objects.update_or_create(
        collect_target_id = collecttarget_object['id'],
        collect_items__reserved_date = target_date_string,
        defaults = {"collect_items": data_json}
    )


def save_platform(data_json):
    """
    플랫폼 저장
    """
    obj = Platform.objects.filter(name=data_json["name"]).first()
    if obj is None:
    # 원래 없는 건 새로 저장
        platform_serializer = PlatformSerializer(data=data_json)
        if platform_serializer.is_valid():
            platform_serializer.save()
    # 있는 건 업데이트
    else:
        platform_serializer = PlatformSerializer(obj, data=data_json)
        if platform_serializer.is_valid():
            platform_serializer.save()


def save_artist(data_json):
    """
    아티스트 저장
    """
    obj = Artist.objects.filter(name=data_json["name"]).first()
    if obj is None:
    # 원래 없는 건 새로 저장
        artist_serializer = ArtistSerializer(data=data_json)
        if artist_serializer.is_valid():
            artist_serializer.save()
    # 있는 건 업데이트
    else:
        artist_serializer = ArtistSerializer(obj, data=data_json)
        if artist_serializer.is_valid():
            artist_serializer.save()


def save_platform_target_item(data_json):
    """
    지금은 사용x
    platform 수집(조사)항목 저장
    """
    obj = PlatformTargetItem.objects.filter(platform=data_json["platform"], target_name=data_json["target_name"],
    xpath=data_json["xpath"]).first()
    if obj is None:
    # 원래 없는 건 새로 저장
        target_item_serializer = PlatformTargetItemSerializer(data=data_json)
        if target_item_serializer.is_valid():
            target_item_serializer.save()
    # 있는 건 업데이트
    else:
        target_item_serializer = PlatformTargetItemSerializer(obj, data=data_json)
        if target_item_serializer.is_valid():
            target_item_serializer.save()


def save_collect_target_item(data_json):
    """
    collect수집(조사)항목 저장
    """
    artist_object = Artist.objects.filter(name=data_json["artist"])
    artist_object = artist_object.values()[0]
    platform_object = Platform.objects.filter(name=data_json["platform"])
    platform_object = platform_object.values()[0]
    collecttarget_object = CollectTarget.objects.filter(artist_id=artist_object["id"], platform_id=platform_object["id"])
    collecttarget_object = collecttarget_object.values()[0]
    obj = CollectTargetItem.objects.filter(collect_target_id=collecttarget_object["id"], target_name=data_json["target_name"]).first()
    data_json = {
        "collect_target": collecttarget_object["id"],
        "target_name": data_json["target_name"],
        "xpath": data_json["xpath"]
    }
    if obj is None:
    # 원래 없는 건 새로 저장
        target_item_serializer = CollectTargetItemSerializer(data=data_json)
        if target_item_serializer.is_valid():
            target_item_serializer.save()
    # 있는 건 업데이트
    else:
        target_item_serializer = CollectTargetItemSerializer(obj, data=data_json)
        if target_item_serializer.is_valid():
            target_item_serializer.save()
    


def save_collect_target(data_json):
    """
    수집대상 저장
    """
    obj = CollectTarget.objects.filter(platform=data_json["platform"], artist=data_json["artist"],
        target_url=data_json["target_url"]).first()
    if obj is None:
    # 원래 없는 건 새로 저장
        collect_target_item_serializer = CollectTargetSerializer(data=data_json)
        if collect_target_item_serializer.is_valid():
            collect_target_item_serializer.save()
    # 있는 건 업데이트
    else:
        collect_target_item_serializer = PlatformTargetItemSerializer(obj, data=data_json)
        if collect_target_item_serializer.is_valid():
            collect_target_item_serializer.save()

def save_schedule(collecttargetid):
    '''
    수집대상의 스케줄 저장
    '''
    schedule_object = Schedule.objects.filter(collect_target_id = collecttargetid).first()
    schedule_data = {
        "collect_target": collecttargetid,
        "schedule_type": "daily",
        "active": True,
        "execute_time": datetime.time(9,0,0)
    }
    schedule_serializer = ScheduleSerializer(schedule_object, data=schedule_data)
    if schedule_serializer.is_valid():
        schedule_serializer.save()
