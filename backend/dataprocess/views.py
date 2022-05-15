import os
from time import time
from django.contrib import auth
from django.shortcuts import render
from account.models import User
from dataprocess.models import CollectData
from config.models import PlatformTargetItem, CollectTargetItem, Schedule
from config.serializers import PlatformTargetItemSerializer, CollectTargetItemSerializer, ScheduleSerializer
from dataprocess.functions import export_datareport, import_datareport, import_collects, import_authinfo
from django.views.decorators.csrf import csrf_exempt
from .resources import *
from .serializers import *
from .models import *
from django.http.response import JsonResponse
from rest_framework.parsers import JSONParser
from rest_framework import status
from django.views.decorators.csrf import csrf_exempt
from utils.api import APIView
from django.shortcuts import render
from django.http import HttpResponse
import datetime
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from .logger import userlogger

# login check using cookie
def logincheck(request):
    # 로그인 정보를 받기 위해 cookie사용
    username = request.COOKIES.get('username')
    if username is not None:
        if User.objects.filter(username=username).exists():
            # 이미 존재하는 username일때만 로그인
            user = User.objects.filter(username=username).first()
            auth.login(request, user)
    return request

# Create your views here.
def base(request):
    '''
    general page
    '''
    platforms = Platform.objects.all() #get all platform info from db
    values = {
        'first_depth' : '데이터 리포트',
        'second_depth': '일별 리포트',
        'platforms': platforms
    }
    request = logincheck(request)
    return render(request, 'dataprocess/daily.html',values)
    
@csrf_exempt
def daily(request):
    if request.method == 'GET':
        '''
        general page
        '''
        platforms = Platform.objects.all() #get all platform info from db
        values = {
            'first_depth' : '데이터 리포트',
            'second_depth': '일별 리포트',
            'platforms': platforms
        }
        request = logincheck(request)
        return render(request, 'dataprocess/daily.html',values)
    else:
        type = request.POST['type']
        if type == 'import':
            '''
            import from excel
            '''
            platforms = Platform.objects.all()  # get all platform info from db
            if not 'importData' in request.FILES:
                values = {
                    'first_depth' : '데이터 리포트',
                    'second_depth': '일별 리포트',
                    'platforms': platforms,
                    'alert': '파일을 첨부해주세요.'
                    }
                request = logincheck(request)
                return render(request, 'dataprocess/daily.html', values)
            import_file = request.FILES['importData']            
            excel_import_date = request.POST.get('excel_import_date', None)  # 0000-0-0 형태

            wb = openpyxl.load_workbook(import_file)
            sheets = wb.sheetnames
            worksheet = wb[sheets[0]]
            import_datareport(worksheet, excel_import_date)

            values = {
                'first_depth' : '데이터 리포트',
                'second_depth': '일별 리포트',
                'platforms': platforms,
                'alert': '저장되었습니다.'
                }
            request = logincheck(request)
            return render(request, 'dataprocess/daily.html',values)
        elif type == 'export':
            '''
            export to excel
            '''
            excel_export_type = request.POST.get('excel_export_days', None) # 누적 or 기간별
            excel_export_start_date = request.POST.get('excel_export_start_date', None) # 0000-0-0 형태
            excel_export_end_date = request.POST.get('excel_export_end_date', None) # 0000-0-0 형태
            book = export_datareport(excel_export_type, excel_export_start_date, excel_export_end_date)
            if excel_export_type == '누적':
                filename = "datareport %s.xlsx" % (excel_export_start_date)
            elif excel_export_type == '기간별':
                filename = "datareport기간별 %s~%s.xlsx" % (excel_export_start_date,excel_export_end_date)
            response = HttpResponse(content=save_virtual_workbook(book), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename='+filename
            return response
        elif type == 'import2':
            '''
            import2 from excel (collect_target_item, artist, platform)
            '''
            platforms = Platform.objects.all()  # get all platform info from db
            if not 'importData' in request.FILES:
                values = {
                    'first_depth' : '데이터 리포트',
                    'second_depth': '일별 리포트',
                    'platforms': platforms,
                    'alert': '파일을 첨부해주세요.'
                    }
                request = logincheck(request)
                return render(request, 'dataprocess/daily.html', values)
            import_file = request.FILES['importData']
            wb = openpyxl.load_workbook(import_file)
            sheets = wb.sheetnames
            worksheet = wb[sheets[0]]
            import_collects(worksheet)
            values = {
                'first_depth' : '데이터 리포트',
                'second_depth': '일별 리포트',
                'platforms': platforms,
                'alert': '저장되었습니다.'
                }
            request = logincheck(request)
            return render(request, 'dataprocess/daily.html',values)
        elif type == 'import3':
            '''
            import3 from excel (auth_info)
            '''
            platforms = Platform.objects.all()  # get all platform info from db
            if not 'importData' in request.FILES:
                values = {
                    'first_depth' : '데이터 리포트',
                    'second_depth': '일별 리포트',
                    'platforms': platforms,
                    'alert': '파일을 첨부해주세요.'
                    }
                request = logincheck(request)
                return render(request, 'dataprocess/daily.html', values)
            import_file = request.FILES['importData']
            wb = openpyxl.load_workbook(import_file)
            sheets = wb.sheetnames
            worksheet = wb[sheets[0]]
            import_authinfo(worksheet)
            values = {
                'first_depth' : '데이터 리포트',
                'second_depth': '일별 리포트',
                'platforms': platforms,
                'alert': '저장되었습니다.'
                }
            request = logincheck(request)
            return render(request, 'dataprocess/daily.html',values)
    
def platform(request):
     '''
     general page
     '''
     values = {
        'first_depth' : '플랫폼 관리',
        'second_depth': '플랫폼 관리'
    }
     request = logincheck(request)
     return render(request, 'dataprocess/platform.html',values)

def artist(request):
    artists = Artist.objects.all()
    values = {
      'first_depth' : '아티스트 관리',
      'second_depth': '데이터 URL 관리',
      'artists': artists
    }
    request = logincheck(request)
    return render(request, 'dataprocess/artist.html',values)

@csrf_exempt
def artist_add(request):
    platforms = Platform.objects.all()
    values = {
      'first_depth' : '아티스트 관리',
      'second_depth': '데이터 URL 관리',
      'platforms' : platforms
    }
    request = logincheck(request)
    return render(request, 'dataprocess/artist_add.html',values)

def login(request):
    values = {
      'first_depth' : '로그인'
    }
    request = logincheck(request)
    return render(request, 'dataprocess/login.html',values)


class PlatformAPI(APIView):
    # @login_required
    def get(self, request):
        '''
        Platform read api
        '''
        try:
            platform_objects = Platform.objects.all()
            if platform_objects.exists():
                platform_objects_values = platform_objects.values()
                platform_datas = []
                for platform_value in platform_objects_values:
                    platform_datas.append(platform_value)
                return JsonResponse(data={'success': True, 'data': platform_datas})
            else:
                return JsonResponse(data={'success': True, 'data': []})
        except:
            return JsonResponse(status=400, data={'success': False})

    # @login_required
    def post(self, request):
        '''
        Platform create api
        '''
        try:
            platform_object = JSONParser().parse(request)
            platform_serializer = PlatformSerializer(data=platform_object)
            if platform_serializer.is_valid():
                # 1. platform 생성
                platform_serializer.save()

                #특정 아티스트 전용 collect_target 생성 시 사용할 코드
                # 2. 현재 존재하는 모든 artist에 대해 collect_target 생성 -> platform과 연결
                artist_objects = Artist.objects.all()
                artist_objects_values = artist_objects.values()
                for artist_objects_value in artist_objects_values:
                    collecttarget = CollectTarget(
                        platform_id = platform_serializer.data['id'],
                        artist_id = artist_objects_value['id']
                        )
                    collecttarget.save()
                    #3. 해당 collecttarget에 대한 schedule 생성
                    schedule_object = Schedule.objects.filter(collect_target_id = collecttarget.id).first()
                    schedule_data = {
                            'collect_target': collecttarget.id,
                            'schedule_type': 'daily',
                            'active': True,
                            'execute_time': datetime.time(9,0,0)
                        }
                    schedule_serializer = ScheduleSerializer(schedule_object, data=schedule_data)
                    if schedule_serializer.is_valid():
                        schedule_serializer.save()
                    #4. 만든 collect target에 대해 수집항목들 생성
                    collecttarget_object = CollectTarget.objects.filter(platform = platform_serializer.data['id'],
                            artist = artist_objects_value['id'])
                    collecttarget_object = collecttarget_object.values()[0]
                    for collect_item in platform_object['collect_items']:
                        collect_item = CollectTargetItem(
                            collect_target_id=collecttarget_object['id'],
                            target_name=collect_item['target_name'],
                            xpath=collect_item['xpath']
                        )
                        collect_item.save()

                return JsonResponse(data={'success': True, 'data': platform_serializer.data}, status=status.HTTP_201_CREATED)
            return JsonResponse(data={'success': False,'data': platform_serializer.errors}, status=400)
        except:
            return JsonResponse(data={'success': False}, status=400)

    # @login_required
    def put(self, request):
        '''
        Platform update api
        '''
        try:
            platform_list = JSONParser().parse(request)
            for platform_object in platform_list:
                platform_data = Platform.objects.filter(pk=platform_object['id']).first()
                if platform_data is None:
                    # 원래 없는 건 새로 저장
                    platform_serializer = PlatformSerializer(data=platform_object)
                    if platform_serializer.is_valid():
                        platform_serializer.save()
                else:
                    data = PlatformSerializer(platform_data).data
                    past_name = data['name']
                    past_url = data['url']
                    cur_name = platform_object['name']
                    cur_url = platform_object['url']
                    platform_serializer = PlatformSerializer(platform_data, data=platform_object)
                    if platform_serializer.is_valid():
                        if past_name != cur_name:
                            userlogger.info(f"[CHANGE]: {past_name} -> {cur_name}")
                        if past_url != cur_url:
                            userlogger.info(f"[CHANGE]: {past_url} -> {cur_url}")
                        platform_serializer.save()
                collecttarget_objects = CollectTarget.objects.filter(platform_id = platform_serializer.data['id'])
                # 해당 platform과 연관된 schedule들 수정 -> artist가 비활성인 애들은 그냥 두고 활성인 애들만 수정
                if collecttarget_objects.exists():
                    collecttarget_values = collecttarget_objects.values()
                    for collecttarget_value in collecttarget_values:
                        artist_object = Artist.objects.get(pk = collecttarget_value['artist_id'])
                        if artist_object.active == True:
                            schedule_objects = Schedule.objects.filter(collect_target_id = collecttarget_value['id'])
                            schedule_objects.update(active = platform_object['active'])
            return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
        except:
            return JsonResponse(data={'success': False}, status=400)

class ArtistAPI(APIView):
    # @login_required
    def get(self, request):
        '''
        Artist read api
        '''
        try:
            artist_objects = Artist.objects.all()
            if artist_objects.exists():
                artist_objects_values = artist_objects.values()
                artist_datas = []
                for artist_value in artist_objects_values:
                    artist_datas.append(artist_value)
                return JsonResponse(status=200,data={'success': True, 'data': artist_datas})
            else:
                return JsonResponse(data={'success': True, 'data': []})
        except:
            return JsonResponse(status=400, data={'success': False})

    # @login_required
    def post(self, request):
        '''
        Artist create api
        '''
        try:
            artist_object = JSONParser().parse(request)
            artist_serializer = ArtistSerializer(data=artist_object)
            if artist_serializer.is_valid():
                # 1. artist 생성
                artist_serializer.save()
                # 2. 현재 존재하는 모든 platform에 대해 collect_target 생성 -> artist와 연결
                for obj in artist_object['urls']:
                    platform_id = Platform.objects.get(name = obj['platform_name']).id
                    artist_id = artist_serializer.data['id']
                    target_url = obj['url1']
                    target_url_2 = obj['url2']
                    collecttarget = CollectTarget(
                        platform_id=platform_id,
                        artist_id=artist_id,
                        target_url=target_url,
                        target_url_2=target_url_2
                    )
                    collecttarget.save()
                    #3. 해당 collecttarget에 대한 schedule 생성(기존 platform의 daily schedule과 똑같이 하기)
                    schedule_object = Schedule.objects.filter(collect_target_id = collecttarget.id).first()
                    execute_time = datetime.time(9,0,0)
                    collecttarget_objects = CollectTarget.objects.filter(artist_id = artist_id)
                    collecttarget_objects = collecttarget_objects.values()
                    for collecttarget_object in collecttarget_objects:
                        schedule_objects = Schedule.objects.filter(schedule_type = 'daily', collect_target_id = collecttarget_object['id']).values()
                        if schedule_objects.exists():
                            execute_time = schedule_objects[0]['execute_time']
                            break
                    schedule_data = {
                            'collect_target': collecttarget.id,
                            'schedule_type': 'daily',
                            'active': True,
                            'execute_time': execute_time
                        }
                    schedule_serializer = ScheduleSerializer(schedule_object, data=schedule_data)
                    if schedule_serializer.is_valid():
                        schedule_serializer.save()
    
                return JsonResponse(data={'success': True, 'data': artist_serializer.data}, status=status.HTTP_201_CREATED)
            return JsonResponse(data={'success': False,'data': artist_serializer.errors}, status=400)
        except:
            return JsonResponse(data={'success': False}, status=400)

    # @login_required
    def put(self, request):
        '''
        Artist update api
        '''
        try:
            artist_list = JSONParser().parse(request)
            for artist_object in artist_list:
                artist_data = Artist.objects.get(id=artist_object["id"])
                data = ArtistSerializer(artist_data).data
                past_name = data['name']
                past_num = data['member_num']
                past_agency = data['agency']
                cur_name = artist_object['name']
                cur_num = artist_object['member_num']
                cur_agency = artist_object['agency']
                artist_serializer = ArtistSerializer(artist_data, data=artist_object)
                if artist_serializer.is_valid():
                    if past_name != cur_name:
                        userlogger.info(f"[CHANGE]: {past_name} -> {cur_name}")
                    if past_num != cur_num:
                        userlogger.info(f"[CHANGE]: {past_num} -> {cur_num}")
                    if past_agency != cur_agency:
                        userlogger.info(f"[CHANGE]: {past_agency} -> {cur_agency}")
                    artist_serializer.save()
                else:
                    return JsonResponse(data={'success': False, 'data': artist_serializer.errors}, status=400)
                collecttarget_objects = CollectTarget.objects.filter(artist_id = artist_serializer.data['id'])
                # 해당 artist와 연관된 schedule들 수정 -> platform이 비활성인 애들은 그냥 두고 활성인 애들만 수정
                if collecttarget_objects.exists():
                    collecttarget_values = collecttarget_objects.values()
                    for collecttarget_value in collecttarget_values:
                        platform_object = Platform.objects.get(pk = collecttarget_value['platform_id'])
                        if platform_object.active == True:
                            schedule_objects = Schedule.objects.filter(collect_target_id = collecttarget_value['id'])
                            schedule_objects.update(active = artist_object['active'])
            return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
        except:
            return JsonResponse(data={'success': False}, status=400)


class PlatformOfArtistAPI(APIView):
    # @login_required
    def get(self, request):
        '''
        Platform of Artist read api
        '''
        try:
            artist = request.GET.get('artist', None)
            # 해당 artist 찾기
            artist_object = Artist.objects.filter(name = artist)
            artist_object = artist_object.values()[0]
            # 해당 artist를 가지는 collect_target들 가져오기
            collecttarget_objects = CollectTarget.objects.filter(artist_id=artist_object['id'])
            if collecttarget_objects.exists():
                collecttarget_objects_values = collecttarget_objects.values()
                platform_datas = []
                for collecttarget_value in collecttarget_objects_values:
                    platform_object = Platform.objects.get(pk=collecttarget_value['platform_id'])
                    platform_datas.append({  
                        'artist_id':artist_object['id'],
                        'platform_id' : collecttarget_value['platform_id'],
                        'id': collecttarget_value['id'],
                        'name': platform_object.name,
                        'target_url':collecttarget_value['target_url'],
                        'target_url_2': collecttarget_value['target_url_2']
                    })
                return JsonResponse(data={'success': True, 'data': platform_datas})
            else:
                return JsonResponse(data={'success': True, 'data': []})
        except:
            return JsonResponse(status=400, data={'success': False})

    # @login_required
    def put(self, request):
        '''
        Platform of Artist update api
        '''
        try:
            collecttarget_list = JSONParser().parse(request)
            data = ''
            for collecttarget_object in collecttarget_list:
                if collecttarget_object['type'] == 'artist-platform-update':
                    CollectTarget.objects.filter(pk=collecttarget_object['id']).update(target_url=collecttarget_object['target_url'])
                    CollectTarget.objects.filter(pk=collecttarget_object['id']).update(target_url_2=collecttarget_object['target_url_2'])
                else:
                    target_obj = CollectTarget.objects.filter(pk=collecttarget_object['id'])
                    target_obj_value = target_obj.values()[0]
                    data = collecttarget_object['new_target_url']

                    if target_obj_value['target_url'] == collecttarget_object['old_target_url']:
                        target_obj.update(target_url = collecttarget_object['new_target_url'])
                    elif target_obj_value['target_url_2'] == collecttarget_object['old_target_url']:
                        target_obj.update(target_url_2 = collecttarget_object['new_target_url'])
                    else:
                        data = ''
            
            return JsonResponse(data={'success': True,'data':data}, status=status.HTTP_201_CREATED)
              
        except:
            return JsonResponse(data={'success': False}, status=400)


class CollectTargetItemAPI(APIView):
    # @login_required
    def get(self, request):
        '''
        CollectTargetItem read api
        '''
        try:
            artist = request.GET.get('artist', None)
            platform = request.GET.get('platform', None)
            # 해당 artist, platform 찾기
            artist_object = Artist.objects.filter(name=artist).first()
            platform_object = Platform.objects.filter(name=platform).first()
            # 해당 artist와 platform을 가지는 collect_target 가져오기
            collecttarget_object = CollectTarget.objects.filter(artist_id=artist_object.id, platform_id=platform_object.id)
            if collecttarget_object.exists():
                collecttarget_object = collecttarget_object.first()
                collecttargetitems_datas = []
                collecttargetitmes_objects = CollectTargetItem.objects.filter(collect_target_id=collecttarget_object.id)
                collecttargetitmes_values = collecttargetitmes_objects.values()
                for collecttargetitmes_value in collecttargetitmes_values:
                    collecttargetitems_datas.append(collecttargetitmes_value)
                # schedule 확인
                schedule_object = Schedule.objects.filter(collect_target_id = collecttarget_object.id)
                if schedule_object.exists():
                    schedule_type = schedule_object.values()[0]['schedule_type']
                else:
                    schedule_type = 'daily'
                return JsonResponse(data={'success': True, 'data': {'items':collecttargetitems_datas, 'schedule_type': schedule_type}})
            else:
                return JsonResponse(data={'success': True, 'data': {'items':[],'schedule_type':'daily'}})
        except:
            return JsonResponse(status=400, data={'success': False})

    # @login_required
    def put(self, request):
        '''
        CollectTargetItem update api
        '''
        try:
            collecttargetitem = JSONParser().parse(request)
            artist = collecttargetitem["artist"]
            platform = collecttargetitem["platform"]
            schedule_type = collecttargetitem["schedule_type"]
            collecttargetitem_list = collecttargetitem['items']
            artist_object = Artist.objects.filter(name = collecttargetitem['artist']).first()
            platform_object = Platform.objects.filter(name = collecttargetitem['platform']).first()
            collecttarget_object = CollectTarget.objects.filter(artist_id=artist_object.id, platform_id=platform_object.id).first()
            for collecttargetitem_object in collecttargetitem_list:
                # 여기 수정!!!!
                collecttargetitem_data = CollectTargetItem.objects.filter(id=collecttargetitem_object['id'],
                                                                          target_name=collecttargetitem_object['target_name'], xpath=collecttargetitem_object['xpath']).first()
                # 없으면 새로 저장
                if collecttargetitem_data is None:
                    collecttargetitem_serializer = CollectTargetItemSerializer(data={
                        'collect_target': collecttarget_object.id,
                        'target_name': collecttargetitem_object['target_name'],
                        'xpath': collecttargetitem_object['xpath']
                    })
                    if collecttargetitem_serializer.is_valid():
                        collecttargetitem_serializer.save()
                    else:
                        # return JsonResponse(data={"success": False, "data": collecttargetitem_serializer.errors}, status=400)
                        return JsonResponse(data={'success': False}, status=400)
                # 있으면 업데이트
                else:
                    collecttargetitem_serializer = CollectTargetItemSerializer(collecttargetitem_data, data=collecttargetitem_object)
                    if collecttargetitem_serializer.is_valid():
                        collecttargetitem_serializer.save()
                        userlogger.debug(f"{artist} - {platform} - {schedule_type}: ")
                    else:
                        return JsonResponse(data={'success': False,'data': collecttargetitem_serializer.errors}, status=400)

            execute_time = datetime.time(9,0,0) #시작 시간
            period = datetime.time(3,0,0) #주기
            collecttarget_objects = CollectTarget.objects.filter(platform_id = platform_object.id)
            collecttarget_objects = collecttarget_objects.values()
            for collecttarget_value in collecttarget_objects:
                schedule_objects = Schedule.objects.filter(schedule_type = collecttargetitem['schedule_type'], collect_target_id = collecttarget_value['id'])
                if schedule_objects.exists():
                    execute_time = schedule_objects.values()[0]['execute_time']
                    period = schedule_objects.values()[0]['period']
                    break
            Schedule.objects.filter(collect_target_id = collecttarget_object.id).update(
                    schedule_type = schedule_type, execute_time = execute_time, period = period)
                
            return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
        except:
            return JsonResponse(data={'success': False}, status=400)

    def delete(self, request):
        '''
        CollectTargetItem delete api
        '''
        try:
            delete_id = JSONParser().parse(request)['id']
            obj = CollectTargetItem.objects.filter(id = delete_id)
            obj.delete()
            return JsonResponse(data={'success': True}, status=status.HTTP_200_OK)
        except:
            return JsonResponse(data={'success': False}, status=400)

#platform collect target API 
class PlatformTargetItemAPI(APIView):
    # @login_required
    def get(self, request):
        '''
        PlatformTargetItem read api
        '''
        try:
            platform = request.GET.get('platform', None)
            # 해당 platform 찾기
            platform_object = Platform.objects.filter(name = platform).first()
            # 해당 platform을 가지는 platform_target 가져오기
            collecttarget_objects = PlatformTargetItem.objects.filter(platform_id = platform_object.id)
            if collecttarget_objects.exists():
                collecttargetitems_datas = []
                collecttarget_objects_value = collecttarget_objects.values()[0]
                collecttargetitmes_objects = PlatformTargetItem.objects.filter(platform_id=collecttarget_objects_value['platform_id'])
                collecttargetitmes_values = collecttargetitmes_objects.values()
                for collecttargetitmes_value in collecttargetitmes_values:
                    collecttargetitems_datas.append(collecttargetitmes_value)
                return JsonResponse(data={'success': True, 'data': collecttargetitems_datas,'platform_id':collecttarget_objects_value['platform_id']})
            else:
                return JsonResponse(data={'success': True, 'data': []})
        except:
            return JsonResponse(status=400, data={'success': False})

    # @login_required
    def put(self, request):
        '''
        PlatformTargetItem update api
        '''
        try:
            collecttargetitem_list = JSONParser().parse(request)
            for i,collecttargetitem_object in enumerate(collecttargetitem_list):
                collecttargetitem_data = PlatformTargetItem.objects.filter(platform_id=collecttargetitem_object['platform'])[i]
                collecttargetitem_serializer = PlatformTargetItemSerializer(collecttargetitem_data, data=collecttargetitem_object)
                if collecttargetitem_serializer.is_valid():
                    collecttargetitem_serializer.save()
                else:
                    return JsonResponse(data={'success': False,'data': collecttargetitem_serializer.errors}, status=400)
            return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
        except:
            return JsonResponse(data={'success': False}, status=400)


class DataReportAPI(APIView):
    def get(self, request):
        '''
        Data-Report read api
        '''
        platform = request.GET.get('platform', None)
        type = request.GET.get('type', None)
        start_date = request.GET.get('start_date', None)
        end_date = request.GET.get('end_date', None)

        # artist name
        # By caching Enhancing Lazy loading Performance 
        artist_objects = list(Artist.objects.filter(active=1))
        artist_list = [query.name for query in artist_objects]

        # platform target names
        # By caching and using Eager laoding Enhancing Lazy Loading Performance
        # Execution time changing: 0.0442s -> 0.0109s
        platform_id = Platform.objects.get(name = platform).id # 가져오고자 하는 platform id
        collect_target_id_list = [collect_target.id for collect_target in list(CollectTarget.objects.filter(platform = platform_id))] # 해당 platfrom에 속한 크롤링 대상 객체들
        collect_target_item = CollectTargetItem.objects.select_related('collect_target').filter(collect_target_id__in=collect_target_id_list) #크롤링 대상 객체
        target_column_list = list(set(item.target_name for item in collect_target_item))
        #platform target names

        #플랫폼 헤더 정보 순서와 db 칼럼 저장 순서 싱크 맞추기
        # Refactoring ambiguous variable names, and apply caching
        # Comments Additional description for variable
        platform_header = []
        crawlered_data_list = [item.collect_items for item in list(CollectData.objects.filter(collect_items__platform=platform))]
        # Configuration of collect_data_dict_list
        # [{'artist':'빅뱅', 'views':215601231 ...}, {'artist':''blabla~', 'views':123123} ...]
        # 크롤링된 데이터의 여러 칼럼을 보여준다.
        
        if len(crawlered_data_list) > 0:
            for key in crawlered_data_list[0].keys():
                if key in target_column_list:
                    platform_header.append(key)
                else:
                    continue
        else:
            platform_header = target_column_list
        

        # try:
        if type == '누적':
            start_date_dateobject = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            start_date_string = start_date_dateobject.strftime("%Y-%m-%d")
            check = False

            # Mainly Improved logic: GET Crawled Data
            # BY Caching, Improve performance of searching data 3.099s -> 0.065s
            # 두 번째로 오래 걸리는 critical path로 약 5배의 성능 상승을 도출함
            for data in crawlered_data_list:
                crawling_artist_list.append(data["artist"])
            
            filter_datas = []
            filter_objects = [collectData.collect_items for collectData in list(CollectData.objects.filter(collect_items__reserved_date=start_date_string, collect_items__platform=platform))]
            for artist in artist_list:
                artist_object = []
                for filter_object in filter_objects:
                    # print(filter_object)
                    if filter_object['artist'] == artist:
                        artist_object.append(filter_object)
                if artist_object:
                    check = True
                    # 같은 날짜에 여러개 있을 때 가장 앞의 것 가져오기
                    filter_value = artist_object[0]
                    filter_datas.append(filter_value)
            
            # 해당날짜에 데이터가 하나라도 있을 때
            if check:
                return JsonResponse(data={'success': True, 'data': filter_datas, 'artists': artist_list, 'platform': platform_header,'crawling_artist_list': crawling_artist_list})
            # 해당날짜에 데이터가 하나도 없을 때
            else:
                crawling_artist_list = []
                objects = CollectData.objects.filter(collect_items__platform=platform)
                objects_value = objects.values()
                for val in objects_value:
                    crawling_artist_list.append(val['collect_items']['artist'])
                return JsonResponse(status=200, data={'success': True, 'data': 'no data', 'artists': artist_list, 'platform': platform_header,
                                                    'crawling_artist_list': crawling_artist_list})

        elif type == '기간별':
            # 전날 값을 구함
            st_time = time()
            start_date_dateobject = datetime.datetime.strptime(start_date, "%Y-%m-%d").date() - datetime.timedelta(1)
            end_date_dateobject = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
            start_date_string = start_date_dateobject.strftime("%Y-%m-%d")
            end_date_string = end_date_dateobject.strftime("%Y-%m-%d")
            check = False

            # Mainly Improved logic: GET Crawled Data
            # BY Caching, Improve performance of searching data 10.102s -> 0.088s
            # 가장 오래 걸리는 부분으로 약 5배의 성능 상승을 도출함

            crawling_artist_list = []
            column_list = ['id', 'artist', 'user_created', 'recorded_date', 'platform', 'url', 'url1', 'url2', 'reserved_date', 'updated_dt']
            filter_objects = [collectData.collect_items for collectData in list(CollectData.objects.filter(collect_items__platform=platform))]
            
            for val in filter_objects:
                crawling_artist_list.append(val['artist'])
            
            filter_datas_total = []
            for artist in artist_list:
                filter_objects_start = []
                filter_objects_end = []
                for object in filter_objects:
                    if object['artist'] == artist:
                        if object['reserved_date'] == start_date_string:
                            filter_objects_start.append(object)
                        if object['reserved_date'] == end_date_string:
                            filter_objects_end.append(object)
                
                # 둘 다 존재할 때
                if filter_objects_start and filter_objects_end:
                    check = True
                    filter_objects_start_value = filter_objects_start[0]

                    # id랑 artist, date 빼고 보내주기
                    data_json = {}
                    filter_objects_end_value = filter_objects_end[0]

                    for field_name in filter_objects_start_value.keys():
                        if field_name not in column_list:
                            if filter_objects_end_value[field_name] is not None and filter_objects_start_value[field_name] is not None:
                                data_json[field_name] = int(filter_objects_end_value[field_name]) - int(filter_objects_start_value[field_name])
                            elif filter_objects_end_value[field_name] is not None:  # 앞의 날짜를 0으로 처리한 형태
                                data_json[field_name] = filter_objects_end_value[field_name]
                            else: # 앞의 날짜가 없다면 0으로 보내기
                                data_json[field_name] = 0
                            data_json[field_name+'_end'] = filter_objects_end_value[field_name]
                        else:  # 숫자 아닌 다른 정보들(user_created 등)
                            data_json[field_name] = filter_objects_start_value[field_name]
                    filter_datas_total.append(data_json)
                elif not filter_objects_start and filter_objects_end:
                    # 시작날짜의 데이터가 존재하지 않고 끝날짜의 데이터만 존재할 때
                    # 시작날짜: 0으로 해서 계산 -> 끝날짜 데이터 자체를 보냄
                    check = True
                    filter_objects_end_value = filter_objects_end[0]
                    filter_datas_total.append(filter_objects_end_value)
            ed_time = time()
            print(ed_time - st_time)
            if check: # 양끝 모두 존재 or 끝날짜만 존재
                return JsonResponse(data={'success': True, 'data': filter_datas_total, 'artists': artist_list, 'platform': platform_header,'crawling_artist_list':crawling_artist_list})
            else: # 끝날짜의 데이터가 아예 존재하지 않을 때
                return JsonResponse(status=400, data={'success': False, 'data': end_date})
        
        else:#누적도 기간별도 아닌 경우(에러처리)
            start_date_dateobject = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            start_date_string = start_date_dateobject.strftime("%Y-%m-%d")
            crawling_artist_list = []
            objects = CollectData.objects.filter(collect_items__platform=platform)
            objects_value = objects.values()
            for val in objects_value:
                crawling_artist_list.append(val['collect_items']['artist'])
            objects = CollectData.objects.filter(collect_items__platform=platform,
                        collect_items__reserved_date = start_date_string)
            if objects.exists():
                platform_queryset_values = objects.values()
                platform_datas = []
                for queryset_value in platform_queryset_values:
                    platform_datas.append(queryset_value['collect_items'])
                return JsonResponse(data={'success': True, 'data': platform_datas, 'artists': artist_list, 'platform': platform_header,'crawling_artist_list':crawling_artist_list})
            else:
                return JsonResponse(status=400, data={'success': False, 'data': start_date})
        # except:
        #     return JsonResponse(status=400, data={'success': False, 'data': start_date})
    
    def post(self, request):
        '''
        Data-Report update api
        '''
        update_data_object = JSONParser().parse(request)
        start_date = update_data_object[len(update_data_object)-1]['start_date']
        platform = update_data_object[len(update_data_object)-1]['platform_name']
        # artist name
        artist_objects = Artist.objects.filter(active=1)

        artist_objects_values = artist_objects.values()
        artist_list = []
        for a in artist_objects_values:
            artist_list.append(a['name'])

        # crawled artist list
        a_objects = CollectData.objects.filter(collect_items__platform=platform)
        a_objects_values = a_objects.values()
        a_list = []
        for val in a_objects_values:
            a_list.append(val['collect_items']['artist'])

        #platform target names
        platform_id = Platform.objects.get(name = platform).id
        collecttargets = CollectTarget.objects.filter(platform = platform_id)
        collecttargets = collecttargets.values()
        platform_list = set()
        for collecttarget in collecttargets:
            platform_objects = CollectTargetItem.objects.filter(collect_target_id = collecttarget['id'])
            platform_objects_values = platform_objects.values()
            for p in platform_objects_values:
                if p['target_name'] in platform_list:
                    continue
                platform_list.add(p['target_name'])
        platform_list = list(platform_list)

        #플랫폼 헤더 정보 순서와 db 칼럼 저장 순서 싱크 맞추기
        platform_header = []
        objects = CollectData.objects.filter(collect_items__platform=platform)
        objects_values = objects.values()
        if len(objects_values) > 0:
            key_list = list(objects_values[0]['collect_items'].keys())
            for key in key_list:
                if key in platform_list:
                    platform_header.append(key)
                else:
                    continue
        else:
            platform_header = platform_list

        try:
            start_date_dateobject = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            start_date_string = start_date_dateobject.strftime("%Y-%m-%d")
            for index, element in enumerate(update_data_object):
                if index == len(update_data_object)-1:
                    break
                artist_object = Artist.objects.filter(name=element['artist'])
                artist_object = artist_object.values()[0]
                platform_object = Platform.objects.filter(name=platform)
                platform_object = platform_object.values()[0]
                collecttarget_object = CollectTarget.objects.filter(platform_id = platform_object['id'],
                                        artist_id = artist_object['id'])
                collecttarget_object = collecttarget_object.values()[0]
                CollectData.objects.update_or_create(
                        collect_target_id = collecttarget_object['id'],
                        collect_items__reserved_date = start_date_string,
                        # collect_items = element,
                        # 바뀌는 값
                        defaults = {'collect_items': element}
                )
                    
            artist_set = set()
            filter_objects = CollectData.objects.filter(collect_items__platform=platform,
                            collect_items__reserved_date = start_date_string)
            if filter_objects.exists():
                filter_objects_values=filter_objects.values()
                filter_datas=[]

                crawling_artist_list = set()
                platform_filter_objects = CollectData.objects.filter(collect_items__platform=platform)
                objects_value = platform_filter_objects.values()
                for val in objects_value:
                    val = val['collect_items']
                    # 각 아티스트가 한번만 들어가도록
                    if val['artist'] in crawling_artist_list:
                        continue
                    crawling_artist_list.add(val['artist'])
                crawling_artist_list = list(crawling_artist_list)
                for filter_value in filter_objects_values:
                    filter_value = filter_value['collect_items']
                    # 각 아티스트당 하나의 데이터만 들어가도록
                    if filter_value['artist'] in artist_set:
                        continue
                    filter_datas.append(filter_value)
                return JsonResponse(data={'success': True, 'data': filter_datas,'artists':artist_list,'platform':platform_header,'crawling_artist_list':crawling_artist_list})
            else:
                # 존재하지 않을 때 -> 플랫폼 전체 데이터를 보자
                filter_objects = CollectData.objects.filter(collect_items__platform=platform)
                crawling_artist_list = set()
                objects_value = filter_objects.values()
                for val in objects_value:
                    val = val['collect_items']
                    # 각 아티스트가 한번만 들어가도록
                    if val['artist'] in crawling_artist_list:
                        continue
                    crawling_artist_list.add(val['artist'])
                crawling_artist_list = list(crawling_artist_list)
                # datename = "%s-%s-%s"%(start_date_dateobject.year, start_date_dateobject.month, start_date_dateobject.day)
                return JsonResponse(status=200, data={'success': True, 'data': 'no data', 'artists': artist_list, 'platform': platform_header, 'crawling_artist_list': crawling_artist_list})
        except:
            return JsonResponse(status=400, data={'success': False})


class ScheduleAPI(APIView):
    def get(self, request):
        '''
        Schedule read api
        '''
        type = request.GET.get('type', None) # 시간별 or 일별
        try:
            if type == '시간별':
                # 해당 플랫폼에 시간별인 아티스트들 가져오기
                platform_objects = Platform.objects.all()
                platform_objects = platform_objects.values()
                hourly_list = []
                for platform_object in platform_objects:
                    period = None
                    execute_time = None
                    hour_artist_list = []
                    collecttarget_objects = CollectTarget.objects.filter(platform_id = platform_object['id'])
                    collecttarget_objects = collecttarget_objects.values()
                    for collecttarget_object in collecttarget_objects:
                        schedule_objects = Schedule.objects.filter(schedule_type = 'hour', collect_target_id = collecttarget_object['id'])
                        if schedule_objects.exists():
                            period = schedule_objects.values()[0]['period']
                            execute_time = schedule_objects.values()[0]['execute_time']
                            artist = Artist.objects.get(pk = collecttarget_object['artist_id'])
                            hour_artist_list.append(artist.name)
                    if period is None:
                        period = datetime.time(0,0,0)
                        execute_time = datetime.time(0,0,0)
                    hourly_list.append({
                        'platform': platform_object['name'],
                        'artists': hour_artist_list,
                        'period': period,
                        'execute_time': execute_time
                    })
                return JsonResponse(data={'success': True, 'data': hourly_list})
            elif type == "일별":
                hourly_list = []
                return JsonResponse(data={'success': True, 'data': hourly_list})
        except:
            return JsonResponse(status=400, data={'success': False})

    def put(self, request):
        '''
        Schedule update api
        '''
        try:
            new_schedule = JSONParser().parse(request)
            schedule_type = new_schedule['schedule_type']
            platform_objects = Platform.objects.filter(name = new_schedule['platform'])
            if platform_objects.exists():
                collecttarget_objects = CollectTarget.objects.filter(platform_id = platform_objects.values()[0]['id'])
                collecttarget_objects = collecttarget_objects.values()
                for collecttarget_object in collecttarget_objects:
                    if schedule_type == 'hour':
                        schedule_objects = Schedule.objects.filter(collect_target_id = collecttarget_object['id'], schedule_type = 'hour')
                        if schedule_objects.exists():
                            schedule_objects.update(period=datetime.time(new_schedule['period'],0,0), execute_time = datetime.time(0,new_schedule['execute_time_minute'],0))
                    elif schedule_type == 'daily':
                        schedule_objects = Schedule.objects.filter(collect_target_id = collecttarget_object['id'], schedule_type = 'daily')
                        if schedule_objects.exists():
                            schedule_objects.update(execute_time = datetime.time(new_schedule['execute_time_hour'],new_schedule['execute_time_minute'],0))
            return JsonResponse(data={'success': True}, status=status.HTTP_201_CREATED)
        except:
            return JsonResponse(data={'success': False}, status=400)
